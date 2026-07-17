#!/usr/bin/env python3
"""
Generate Excel file for tournament participants
Grouped by session with Team Name and Nickname only
"""

import json
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Paths
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE_DIR, 'database', 'db.json')
OUTPUT_DIR = os.path.join(BASE_DIR, 'download')


def load_db():
    """Load database from db.json"""
    try:
        with open(DB_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading DB: {e}")
        return None


def get_today_sessions(db):
    """Get today's sessions data"""
    today = datetime.now().strftime('%Y-%m-%d')
    sessions_data = db.get('sessions', {}).get(today, [])
    return sessions_data


def generate_excel():
    """Generate Excel file with participant data grouped by session"""
    
    # Load database
    db = load_db()
    if not db:
        return None
    
    # Get config
    config = db.get('config', {})
    max_slot = config.get('max_slot_per_session', 12)
    prefix_kode = config.get('prefix_kode', 'FTSG')
    
    # Get today's date for filename
    today_str = datetime.now().strftime('%Y-%m-%d')
    today_date = datetime.now().strftime('%d/%m/%Y')
    
    # Get all participants (for today)
    today_sessions = get_today_sessions(db)
    
    # Group by session
    sessions = {}
    for p in today_sessions:
        sesi = p.get('session', 'Unknown')
        if sesi not in sessions:
            sessions[sesi] = []
        sessions[sesi].append(p)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Peserta"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    cell_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # Title
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = f"DATA PESERTA TURNAMEN - {today_date}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center_align
    
    # Subtitle with total
    total_peserta = len(today_sessions)
    ws.merge_cells('A2:D2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Total Peserta: {total_peserta}"
    subtitle_cell.font = Font(italic=True, size=10)
    subtitle_cell.alignment = center_align
    
    current_row = 4
    
    # Sort sessions by number
    def get_session_num(sesi_name):
        try:
            return int(sesi_name.replace('Sesi ', ''))
        except:
            return 999
    
    sorted_sessions = sorted(sessions.keys(), key=get_session_num)
    
    # Write each session
    for sesi_name in sorted_sessions:
        peserta_list = sessions[sesi_name]
        
        # Session header
        ws.merge_cells(f'A{current_row}:D{current_row}')
        session_header = ws[f'A{current_row}']
        session_header.value = f"{sesi_name} ({len(peserta_list)}/{max_slot} slot)"
        session_header.font = Font(bold=True, size=11, color="2F5496")
        session_header.fill = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
        current_row += 1
        
        # Column headers
        headers = ['No', 'Nama Team', 'Nickname', 'Kode']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border
        current_row += 1
        
        # Participant data
        for idx, peserta in enumerate(peserta_list, 1):
            data = [
                idx,
                peserta.get('team', '-'),
                peserta.get('kapten', '-'),
                peserta.get('id', '-')
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = cell_border
                if col == 1:  # No column
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
            
            current_row += 1
        
        # Empty slots (if any)
        empty_slots = max_slot - len(peserta_list)
        if empty_slots > 0:
            for i in range(empty_slots):
                data = [len(peserta_list) + i + 1, '-', '-', '-']
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.border = cell_border
                    cell.alignment = center_align if col == 1 else left_align
                    # Light gray for empty slots
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                current_row += 1
        
        # Space between sessions
        current_row += 1
    
    # If no participants
    if not sorted_sessions:
        ws.merge_cells('A4:D4')
        no_data = ws['A4']
        no_data.value = "Belum ada peserta hari ini"
        no_data.font = Font(italic=True, color="808080")
        no_data.alignment = center_align
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 6   # No
    ws.column_dimensions['B'].width = 25  # Nama Team
    ws.column_dimensions['C'].width = 20  # Nickname
    ws.column_dimensions['D'].width = 15  # Kode
    
    # Ensure output directory exists
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    # Save file
    filename = f"data_peserta_{today_str}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    
    try:
        wb.save(filepath)
        print(f"✅ Excel generated: {filepath}")
        return filepath
    except Exception as e:
        print(f"❌ Error saving Excel: {e}")
        return None


if __name__ == "__main__":
    result = generate_excel()
    if result:
        print(f"File: {result}")
    else:
        sys.exit(1)
